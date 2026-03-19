"""
nifty100_tracker.py — Ticker-Teller Daily Automation  (v2 — Smart Retry)
=========================================================================
Runs daily predictions for all NIFTY 100 companies and records results
into individual Excel workbooks with conditional formatting.

New in v2
─────────
  Auto symbol correction
      When a ticker returns HTTP 404 / "not found" / "possibly delisted",
      the script queries the Yahoo Finance search API to find the correct
      current symbol, patches the in-memory list, and retries automatically.
      Every confirmed fix is saved to  state/ticker_fixes.json  so it
      persists across runs without touching the source file.

  Smart retry (packet-resend logic)
      Transient failures (network timeout, connection reset, rate-limit)
      trigger an exponential-backoff retry loop — just like TCP retransmission.
      Symbol-not-found errors are never retried blindly; they go straight to
      the web-search correction path instead.

Usage:
    python nifty100_tracker.py               # Run full 30-day automation
    python nifty100_tracker.py --once        # Run a single day immediately
    python nifty100_tracker.py --resume      # Resume from last checkpoint
    python nifty100_tracker.py --tickers RELIANCE.NS TCS.NS

Directory layout (auto-created):
    workbooks/          <- one .xlsx per company
    logs/               <- daily + master log files
    state/              <- checkpoint.json + ticker_fixes.json
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import requests
import torch
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sklearn.preprocessing import MinMaxScaler

# Suppress noisy Streamlit warnings when running headless
logging.getLogger("streamlit").setLevel(logging.ERROR)

# -- Path bootstrap so src/ imports work from any CWD -------------------------
ROOT_DIR = Path(__file__).parent
sys.path.insert(0, str(ROOT_DIR))

from src.data import (
    FACTOR_FEATURE_COLS, FEATURE_COLS, N_FACTOR_FEATURES,
    N_PRICE_FEATURES, PRICE_FEATURE_COLS,
    add_quant_features, add_technical_features,
    build_sequences, fetch_data, fetch_market_data,
    returns_to_prices, split_data,
)
from src.model import (
    make_loaders, monte_carlo_forecast, predict_test_set, train_model,
)

# =============================================================================
# CONFIGURATION
# =============================================================================

WORKBOOK_DIR = ROOT_DIR / "workbooks"
LOG_DIR      = ROOT_DIR / "logs"
STATE_DIR    = ROOT_DIR / "state"

TOTAL_DAYS = 30
RUN_HOUR   = 18       # 6 PM IST -- after NSE market close (15:30 IST)
RUN_MINUTE = 0

# -- Retry configuration ------------------------------------------------------
# Transient errors  -> exponential backoff  (like TCP retransmission)
TRANSIENT_MAX_RETRIES = 5
TRANSIENT_BASE_DELAY  = 30    # seconds; doubles each attempt: 30,60,120,240,480s

# Model hyper-parameters
MODEL_CFG = dict(
    period        = 730,
    seq_length    = 60,
    forecast_days = 1,
    epochs        = 50,
    patience      = 10,
    hidden_size   = 128,
    cnn_channels  = 64,
    num_heads     = 4,
    dropout       = 0.2,
)

# -- Excel visual theme -------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
SUB_FILL    = PatternFill("solid", fgColor="0F2744")
HEADER_FONT = Font(bold=True, color="E2E8F0", name="Arial", size=10)
SUB_FONT    = Font(bold=True, color="94A3B8", name="Arial", size=9)
DATA_FONT   = Font(name="Arial", size=9)
MONO_FONT   = Font(name="Courier New", size=8, color="94A3B8")
THIN_BORDER = Border(
    left   = Side(style="thin", color="1E3A5F"),
    right  = Side(style="thin", color="1E3A5F"),
    top    = Side(style="thin", color="1E3A5F"),
    bottom = Side(style="thin", color="1E3A5F"),
)

# =============================================================================
# NIFTY 100 TICKERS  (Yahoo Finance .NS format)
# =============================================================================

NIFTY100_TICKERS: list[str] = [
    "RELIANCE.NS","TCS.NS","HDFCBANK.NS","BHARTIARTL.NS","ICICIBANK.NS",
    "INFY.NS","SBIN.NS","HINDUNILVR.NS","ITC.NS","LT.NS",
    "BAJFINANCE.NS","HCLTECH.NS","MARUTI.NS","SUNPHARMA.NS","ADANIENT.NS",
    "KOTAKBANK.NS","TITAN.NS","ONGC.NS","NTPC.NS","POWERGRID.NS",
    "ULTRACEMCO.NS","WIPRO.NS","BAJAJFINSV.NS","ADANIPORTS.NS","NESTLEIND.NS",
    "JSWSTEEL.NS","TATASTEEL.NS","HINDALCO.NS","DIVISLAB.NS","DRREDDY.NS",
    "CIPLA.NS","ASIANPAINT.NS","COALINDIA.NS","BPCL.NS","GRASIM.NS",
    "EICHERMOT.NS","HEROMOTOCO.NS","APOLLOHOSP.NS","BRITANNIA.NS","TATACONSUM.NS",
    "M&M.NS","SBILIFE.NS","HDFCLIFE.NS","SHRIRAMFIN.NS","BAJAJ-AUTO.NS",
    "INDUSINDBK.NS","TECHM.NS","TATAMOTORS.NS","AXISBANK.NS","LTF.NS",
    "LTIM.NS","PIDILITIND.NS","DMART.NS","HAVELLS.NS","SIEMENS.NS",
    "MUTHOOTFIN.NS","BEL.NS","GODREJCP.NS","BERGEPAINT.NS","CHOLAFIN.NS",
    "TRENT.NS","COLPAL.NS","AMBUJACEM.NS","TORNTPHARM.NS","DABUR.NS",
    "SRF.NS","MARICO.NS","BANKBARODA.NS","NAUKRI.NS","ETERNAL.NS",
    "ICICIPRULI.NS","JIOFIN.NS","INDIGO.NS","PNB.NS","ADANIGREEN.NS",
    "ADANIENSOL.NS","ADANIPOWER.NS","ABB.NS","MOTHERSON.NS","PGHH.NS",
    "MAXHEALTH.NS","OBEROIRLTY.NS","MFSL.NS","PERSISTENT.NS","POLYCAB.NS",
    "CANBK.NS","INDIANB.NS","PETRONET.NS","GAIL.NS","IOC.NS",
    "RECLTD.NS","PFC.NS","NHPC.NS","IRFC.NS","HUDCO.NS",
    "LINDEINDIA.NS","CGPOWER.NS","PHOENIXLTD.NS","TIINDIA.NS","NYKAA.NS",
]

# =============================================================================
# ERROR CLASSIFICATION
# Two distinct failure modes require two distinct responses.
# =============================================================================

_NOT_FOUND_PHRASES = (
    "not found",
    "possibly delisted",
    "no timezone found",
    "404",
    "symbol may be delisted",
    "quote not found",
    "no data found",
    "delisted",
    "no ohlcv data",           # raised by run_prediction() itself
)

_TRANSIENT_PHRASES = (
    "connection",
    "timeout",
    "timed out",
    "reset",
    "refused",
    "temporary",
    "rate limit",
    "too many requests",
    "service unavailable",
    "502", "503", "504",
    "read error",
    "ssl",
    "eof",
    "remote end closed",
    "chunkedencodingerror",
)


def is_symbol_not_found(exc: Exception) -> bool:
    """True when the error clearly means the ticker symbol is wrong / gone."""
    msg = str(exc).lower()
    return any(p in msg for p in _NOT_FOUND_PHRASES)


def is_transient_error(exc: Exception) -> bool:
    """True for recoverable network / API errors that are worth retrying."""
    msg = str(exc).lower()
    return any(p in msg for p in _TRANSIENT_PHRASES)


# =============================================================================
# TICKER FIX REGISTRY  (persisted to state/ticker_fixes.json)
# =============================================================================

FIXES_FILE = STATE_DIR / "ticker_fixes.json"


def load_fixes() -> dict[str, str]:
    """Load previously discovered symbol corrections  {old_symbol: new_symbol}."""
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    if FIXES_FILE.exists():
        with open(FIXES_FILE) as f:
            return json.load(f)
    return {}


def save_fixes(fixes: dict[str, str]) -> None:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    with open(FIXES_FILE, "w") as f:
        json.dump(fixes, f, indent=2)


# =============================================================================
# WEB SEARCH -- AUTO SYMBOL CORRECTION
# Queries Yahoo Finance's own search API (no third-party key needed).
# =============================================================================

_YF_SEARCH_URL = "https://query2.finance.yahoo.com/v1/finance/search"
_YF_HEADERS    = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    ),
    "Accept": "application/json",
}


def _yahoo_search(query: str, logger: logging.Logger) -> list[dict]:
    """Hit the Yahoo Finance search endpoint and return the quotes list."""
    try:
        resp = requests.get(
            _YF_SEARCH_URL,
            params  = {"q": query, "quotesCount": 10, "newsCount": 0, "listsCount": 0},
            headers = _YF_HEADERS,
            timeout = 10,
        )
        resp.raise_for_status()
        return resp.json().get("quotes", [])
    except Exception as exc:
        logger.warning(f"Yahoo Finance search API error for '{query}': {exc}")
        return []


def search_correct_ticker(
    bad_ticker: str,
    logger:     logging.Logger,
) -> Optional[str]:
    """
    Given a failing ticker (e.g. 'ADANITRANS.NS'), search Yahoo Finance
    to find its current correct symbol.

    Strategy
    --------
    1. Strip .NS, search by the stem word(s).
    2. Filter to NSE equities only (.NS suffix).
    3. Score candidates by overlap with the original stem.
    4. Probe the top candidate with a tiny yfinance download to confirm it works.
    5. Return the confirmed symbol, or None if nothing useful found.
    """
    stem = (
        bad_ticker
        .replace(".NS", "")
        .replace("-", " ")
        .replace("&", "AND")
    )
    logger.info(
        f"[{bad_ticker}] Querying Yahoo Finance search API  (query: '{stem}')..."
    )

    quotes = _yahoo_search(stem, logger)

    # Filter to NSE equities only
    nse = [
        q for q in quotes
        if str(q.get("symbol", "")).endswith(".NS")
        and q.get("typeDisp", "").upper() in ("EQUITY", "")
    ]

    # Broaden to first word if initial search returned nothing
    if not nse:
        first_word = stem.split()[0]
        if first_word != stem:
            logger.info(
                f"[{bad_ticker}] No results for '{stem}' — "
                f"broadening search to '{first_word}'..."
            )
            quotes = _yahoo_search(first_word, logger)
            nse    = [
                q for q in quotes
                if str(q.get("symbol", "")).endswith(".NS")
                and q.get("typeDisp", "").upper() in ("EQUITY", "")
            ]

    if not nse:
        logger.warning(f"[{bad_ticker}] Web search found no NSE candidates.")
        return None

    # Score by stem overlap
    stem_upper = stem.upper().replace(" ", "")

    def _score(q: dict) -> int:
        sym    = q.get("symbol", "").replace(".NS", "").upper()
        common = sum(1 for a, b in zip(sym, stem_upper) if a == b)
        sub    = 10 if (stem_upper in sym or sym in stem_upper) else 0
        exact  = 100 if sym == stem_upper else 0
        return exact + common + sub

    nse.sort(key=_score, reverse=True)
    best         = nse[0]
    new_symbol   = best["symbol"]
    display_name = best.get("shortname") or best.get("longname") or new_symbol

    logger.info(
        f"[{bad_ticker}] Top candidate: {new_symbol}  ({display_name})  "
        f"score={_score(best)}"
    )

    # Probe: confirm the candidate actually returns data before committing
    import yfinance as yf
    try:
        probe = yf.download(new_symbol, period="5d", progress=False)
        if probe is None or probe.empty:
            logger.warning(
                f"[{bad_ticker}] Probe download for {new_symbol} returned no data "
                f"— discarding this candidate."
            )
            return None
    except Exception as exc:
        logger.warning(
            f"[{bad_ticker}] Probe download for {new_symbol} failed: {exc}"
        )
        return None

    logger.info(
        f"[{bad_ticker}] Confirmed replacement: {bad_ticker} --> {new_symbol}"
    )
    return new_symbol


# =============================================================================
# LOGGING
# =============================================================================

def setup_logging(run_date: str) -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"tracker_{run_date}.log"
    master   = LOG_DIR / "master.log"

    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    logger = logging.getLogger("nifty100_tracker")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    mh = logging.FileHandler(master, encoding="utf-8")
    mh.setLevel(logging.INFO)
    mh.setFormatter(fmt)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(mh)
    logger.addHandler(ch)
    return logger


# =============================================================================
# STATE / CHECKPOINT
# =============================================================================

STATE_FILE = STATE_DIR / "checkpoint.json"


def load_state() -> dict:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    if STATE_FILE.exists():
        with open(STATE_FILE) as f:
            return json.load(f)
    return {"days_completed": 0, "start_date": None, "daily_records": {}}


def save_state(state: dict) -> None:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2, default=str)


def mark_ticker_done(state: dict, run_date: str, ticker: str) -> None:
    state.setdefault("daily_records", {}).setdefault(run_date, [])
    if ticker not in state["daily_records"][run_date]:
        state["daily_records"][run_date].append(ticker)
    save_state(state)


def is_ticker_done(state: dict, run_date: str, ticker: str) -> bool:
    return ticker in state.get("daily_records", {}).get(run_date, [])


# =============================================================================
# EXCEL WORKBOOK HELPERS
# =============================================================================

COL_DATE     = 1   # A
COL_TICKER   = 2   # B
COL_PRICE_F  = 3   # C
COL_FACTOR_F = 4   # D
COL_ACTUAL   = 5   # E
COL_PRED     = 6   # F
COL_ERROR    = 7   # G
COL_SIGNED   = 8   # H  (hidden -- drives colour logic)

HEADER_ROW = 1
SUB_ROW    = 2
DATA_START = 3


def _safe_filename(ticker: str) -> str:
    return ticker.replace(".", "_").replace("/", "_").replace("&", "AND")


def _workbook_path(ticker: str) -> Path:
    WORKBOOK_DIR.mkdir(parents=True, exist_ok=True)
    return WORKBOOK_DIR / f"{_safe_filename(ticker)}.xlsx"


def _apply_header_border(cell, fill=None, font=None, align_center=True) -> None:
    if fill: cell.fill = fill
    if font: cell.font = font
    cell.border = THIN_BORDER
    if align_center:
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )


def create_workbook(ticker: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"
    ws.sheet_properties.tabColor = "1E3A5F"

    ws.merge_cells(start_row=1, start_column=COL_DATE, end_row=2, end_column=COL_DATE)
    _apply_header_border(ws.cell(1, COL_DATE, "Date"), HEADER_FILL, HEADER_FONT)

    ws.merge_cells(start_row=1, start_column=COL_TICKER, end_row=2, end_column=COL_TICKER)
    _apply_header_border(ws.cell(1, COL_TICKER, "Ticker"), HEADER_FILL, HEADER_FONT)

    ws.merge_cells(start_row=1, start_column=COL_PRICE_F,
                   end_row=1, end_column=COL_FACTOR_F)
    _apply_header_border(ws.cell(1, COL_PRICE_F, "Model Input"), HEADER_FILL, HEADER_FONT)

    c = ws.cell(SUB_ROW, COL_PRICE_F,
                f"Price Features ({N_PRICE_FEATURES})\n" + " . ".join(PRICE_FEATURE_COLS))
    _apply_header_border(c, SUB_FILL, SUB_FONT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    c = ws.cell(SUB_ROW, COL_FACTOR_F,
                f"Quant Factor Features ({N_FACTOR_FEATURES})\n"
                + " . ".join(FACTOR_FEATURE_COLS))
    _apply_header_border(c, SUB_FILL, SUB_FONT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.merge_cells(start_row=1, start_column=COL_ACTUAL, end_row=2, end_column=COL_ACTUAL)
    _apply_header_border(
        ws.cell(1, COL_ACTUAL, "Actual Price\n(Rs)"), HEADER_FILL, HEADER_FONT
    )

    ws.merge_cells(start_row=1, start_column=COL_PRED, end_row=2, end_column=COL_PRED)
    _apply_header_border(
        ws.cell(1, COL_PRED, "Predicted Price\n(Rs)"), HEADER_FILL, HEADER_FONT
    )

    ws.merge_cells(start_row=1, start_column=COL_ERROR, end_row=2, end_column=COL_ERROR)
    _apply_header_border(
        ws.cell(1, COL_ERROR, "Error (%)\n|((Actual-Pred)/Actual)x100|"),
        HEADER_FILL, HEADER_FONT,
    )

    ws.column_dimensions[get_column_letter(COL_DATE)].width     = 14
    ws.column_dimensions[get_column_letter(COL_TICKER)].width   = 14
    ws.column_dimensions[get_column_letter(COL_PRICE_F)].width  = 60
    ws.column_dimensions[get_column_letter(COL_FACTOR_F)].width = 50
    ws.column_dimensions[get_column_letter(COL_ACTUAL)].width   = 16
    ws.column_dimensions[get_column_letter(COL_PRED)].width     = 16
    ws.column_dimensions[get_column_letter(COL_ERROR)].width    = 22

    ws.row_dimensions[HEADER_ROW].height = 28
    ws.row_dimensions[SUB_ROW].height    = 70
    ws.freeze_panes                      = ws.cell(DATA_START, 1)

    ts = wb.create_sheet("Info")
    company = ticker.replace(".NS", "").replace("-", " ")
    ts["A1"] = f"Ticker-Teller  --  {company}"
    ts["A1"].font = Font(bold=True, size=14, color="38BDF8", name="Arial")
    ts["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ts["A2"].font = Font(size=9, color="94A3B8", name="Arial")
    ts["A3"] = "Model: CNN -> BiLSTM -> Multi-Head Attention"
    ts["A4"] = f"Price features  ({N_PRICE_FEATURES}): " + ", ".join(PRICE_FEATURE_COLS)
    ts["A5"] = f"Factor features ({N_FACTOR_FEATURES}): " + ", ".join(FACTOR_FEATURE_COLS)
    ts["A6"] = "Target: log return = log(Close[t] / Close[t-1])"
    ts["A7"] = "Error formula: ABS(((Actual - Predicted) / Actual) * 100)"
    ts["A8"] = "GREEN = model underestimated; RED = model overestimated"
    for r in range(1, 9):
        ts.cell(r, 1).alignment = Alignment(wrap_text=True)
    ts.column_dimensions["A"].width = 90

    wb.save(_workbook_path(ticker))


def append_prediction_row(
    ticker:       str,
    run_date:     str,
    actual:       float,
    predicted:    float,
    price_feats:  str,
    factor_feats: str,
) -> None:
    path = _workbook_path(ticker)
    if not path.exists():
        create_workbook(ticker)

    wb       = load_workbook(path)
    ws       = wb["Predictions"]
    next_row = max(ws.max_row + 1, DATA_START)

    e_col = get_column_letter(COL_ACTUAL)
    f_col = get_column_letter(COL_PRED)
    h_col = get_column_letter(COL_SIGNED)

    ws.cell(next_row, COL_DATE,     run_date).number_format = "YYYY-MM-DD"
    ws.cell(next_row, COL_TICKER,   ticker)
    ws.cell(next_row, COL_PRICE_F,  price_feats)
    ws.cell(next_row, COL_FACTOR_F, factor_feats)
    ws.cell(next_row, COL_ACTUAL,   actual).number_format    = '#,##0.00'
    ws.cell(next_row, COL_PRED,     predicted).number_format = '#,##0.00'

    ws.cell(next_row, COL_ERROR,
            f'=ABS((({e_col}{next_row}-{f_col}{next_row})/{e_col}{next_row})*100)'
            ).number_format = '0.00"%"'

    ws.cell(next_row, COL_SIGNED,
            f'=(({e_col}{next_row}-{f_col}{next_row})/{e_col}{next_row})*100'
            ).number_format = '0.00"%"'

    for col in range(1, COL_SIGNED + 1):
        cell        = ws.cell(next_row, col)
        cell.border = THIN_BORDER
        if col in (COL_PRICE_F, COL_FACTOR_F):
            cell.font      = MONO_FONT
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
        else:
            cell.font      = DATA_FONT
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )

    ws.row_dimensions[next_row].height = 14

    raw_signed = ((actual - predicted) / actual * 100) if actual != 0 else 0
    ws.cell(next_row, COL_ERROR).fill = (
        PatternFill("solid", fgColor="052E16") if raw_signed >= 0
        else PatternFill("solid", fgColor="450A0A")
    )
    ws.cell(next_row, COL_ERROR).font = Font(
        color  = "4ADE80" if raw_signed >= 0 else "FCA5A5",
        bold   = True,
        name   = "Arial",
        size   = 9,
    )
    ws.column_dimensions[h_col].hidden = True
    wb.save(path)


# =============================================================================
# FEATURE SNAPSHOT
# =============================================================================

def _feature_snapshot(df: pd.DataFrame, feats: list[str]) -> str:
    last = df[feats].dropna().iloc[-1]
    return " | ".join(f"{k}={v:.4f}" for k, v in last.items())


# =============================================================================
# CORE PREDICTION LOGIC
# =============================================================================

def run_prediction(ticker: str, logger: logging.Logger) -> Optional[dict]:
    """
    Full Ticker-Teller pipeline for one ticker.
    Raises ValueError / Exception on any failure so the smart-retry
    wrapper can classify it correctly.
    """
    cfg    = MODEL_CFG
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    logger.debug(f"[{ticker}] Fetching OHLCV...")
    df_raw = fetch_data(ticker, cfg["period"])
    if df_raw is None or df_raw.empty:
        raise ValueError(
            f"No OHLCV data returned for {ticker} -- possibly delisted"
        )

    logger.debug(f"[{ticker}] Fetching SPY market data...")
    market_df = fetch_market_data(cfg["period"])
    if market_df is None or market_df.empty:
        logger.warning(f"[{ticker}] No SPY data -- using Close proxy.")
        market_df = df_raw[["Close"]].copy()

    df = add_technical_features(df_raw)
    df = add_quant_features(df, market_df)
    df = df.dropna(subset=FEATURE_COLS)[FEATURE_COLS]

    if len(df) < cfg["seq_length"] + 10:
        raise ValueError(
            f"Insufficient rows ({len(df)}) after feature engineering for {ticker}"
        )

    price_feats_str  = _feature_snapshot(df, PRICE_FEATURE_COLS)
    factor_feats_str = _feature_snapshot(df, FACTOR_FEATURE_COLS)

    scaler      = MinMaxScaler(feature_range=(0, 1))
    scaled_data = scaler.fit_transform(df)
    raw_close   = df["Close"].values
    X, y        = build_sequences(scaled_data, raw_close, cfg["seq_length"])

    (X_train, y_train), (X_val, y_val), (X_test, y_test), split, val_split = \
        split_data(X, y)

    train_loader, val_loader, test_loader = make_loaders(
        X_train, y_train, X_val, y_val, X_test, y_test
    )

    import unittest.mock as mock
    dummy = mock.MagicMock()
    with mock.patch("streamlit.progress", return_value=dummy), \
         mock.patch("streamlit.empty",    return_value=dummy):
        logger.debug(f"[{ticker}] Training model...")
        model, _ = train_model(
            train_loader, val_loader,
            input_size   = len(FEATURE_COLS),
            epochs       = cfg["epochs"],
            patience     = cfg["patience"],
            device       = device,
            hidden_size  = cfg["hidden_size"],
            dropout      = cfg["dropout"],
            cnn_channels = cfg["cnn_channels"],
            num_heads    = cfg["num_heads"],
        )

    last_close = float(df_raw["Close"].iloc[-1])
    mc_raw     = monte_carlo_forecast(
        model, scaled_data,
        raw_close_last = last_close,
        scaler         = scaler,
        n_features     = len(FEATURE_COLS),
        seq_length     = cfg["seq_length"],
        forecast_days  = 1,
        device         = device,
    )
    predicted_price = float(mc_raw[0]["mean"])

    logger.info(
        f"[{ticker}] actual={last_close:.2f}  predicted={predicted_price:.2f}  "
        f"error={abs((last_close - predicted_price) / last_close * 100):.2f}%"
    )
    return dict(
        actual        = last_close,
        predicted     = predicted_price,
        price_feats   = price_feats_str,
        factor_feats  = factor_feats_str,
    )


# =============================================================================
# SMART FETCH  --  AUTO CORRECTION + EXPONENTIAL BACKOFF
#
#   Two completely separate failure modes, two completely separate responses:
#
#   MODE A  --  Symbol not found (404 / delisted / no timezone)
#   ─────────────────────────────────────────────────────────────
#   Step 1 : Query Yahoo Finance search API for the correct current symbol.
#   Step 2 : Probe the top candidate with a 5-day download to confirm.
#   Step 3 : Save the fix to state/ticker_fixes.json (persists across runs).
#   Step 4 : Retry once with the corrected symbol.
#   Step 5 : If still failing -> log as unresolvable, skip for today.
#
#   MODE B  --  Transient error (timeout, 502, connection reset ...)
#   ─────────────────────────────────────────────────────────────────
#   Exponential backoff: wait 30s -> 60s -> 120s -> 240s -> 480s
#   between retries.  Mirrors TCP retransmission -- keep re-sending
#   until the packet lands or the max attempts are exhausted.
# =============================================================================

def fetch_with_smart_retry(
    ticker:   str,
    fixes:    dict[str, str],
    state:    dict,
    run_date: str,
    logger:   logging.Logger,
) -> tuple[Optional[dict], str]:
    """
    Returns (result_dict | None, effective_ticker_used).
    """
    effective_ticker = fixes.get(ticker, ticker)
    if effective_ticker != ticker:
        logger.info(f"[{ticker}] Applying known fix -> {effective_ticker}")

    attempt         = 0
    symbol_searched = False   # search at most once per ticker per run

    while True:
        attempt += 1
        try:
            result = run_prediction(effective_ticker, logger)

            # Persist any new fix that was just confirmed
            if effective_ticker != ticker and ticker not in fixes:
                fixes[ticker] = effective_ticker
                save_fixes(fixes)

            return result, effective_ticker

        except Exception as exc:

            # ------------------------------------------------------------------
            # MODE A: symbol is wrong -- search for the correct one
            # ------------------------------------------------------------------
            if is_symbol_not_found(exc) and not symbol_searched:
                logger.warning(
                    f"[{effective_ticker}] Symbol not found  -->  "
                    f"searching Yahoo Finance for correct symbol..."
                )
                symbol_searched = True
                new_symbol = search_correct_ticker(effective_ticker, logger)

                if new_symbol and new_symbol != effective_ticker:
                    fixes[ticker]    = new_symbol
                    effective_ticker = new_symbol
                    save_fixes(fixes)
                    logger.info(
                        f"[{ticker}] Retrying with corrected symbol: {new_symbol}"
                    )
                    continue   # immediate retry with new symbol; no sleep needed

                else:
                    logger.error(
                        f"[{ticker}] Web search could not find a valid replacement. "
                        f"Skipping for today."
                    )
                    return None, effective_ticker

            # Searched but still failing
            if is_symbol_not_found(exc) and symbol_searched:
                logger.error(
                    f"[{ticker}] Still not found after symbol correction. Skipping."
                )
                return None, effective_ticker

            # ------------------------------------------------------------------
            # MODE B: transient network / API error -- exponential backoff
            # ------------------------------------------------------------------
            if attempt > TRANSIENT_MAX_RETRIES:
                logger.error(
                    f"[{effective_ticker}] Exhausted {TRANSIENT_MAX_RETRIES} retries. "
                    f"Last error: {str(exc)[:120]}"
                )
                return None, effective_ticker

            delay = TRANSIENT_BASE_DELAY * (2 ** (attempt - 1))
            logger.warning(
                f"[{effective_ticker}] Attempt {attempt}/{TRANSIENT_MAX_RETRIES} "
                f"failed  ({str(exc)[:80]})  -->  "
                f"retrying in {delay}s  [backoff x2 per attempt]"
            )
            time.sleep(delay)


# =============================================================================
# DAILY RUN
# =============================================================================

def run_daily(tickers: list[str], state: dict, logger: logging.Logger) -> None:
    run_date = datetime.now().strftime("%Y-%m-%d")
    logger.info(f"=== Daily run  {run_date} ===  ({len(tickers)} tickers)")

    fixes = load_fixes()
    if fixes:
        logger.info(
            f"Loaded {len(fixes)} known ticker fix(es) from ticker_fixes.json"
        )

    success, skipped, corrected, failed = 0, 0, 0, 0

    for i, ticker in enumerate(tickers, 1):

        # Check if already done today (handles both old and corrected symbols)
        effective = fixes.get(ticker, ticker)
        if is_ticker_done(state, run_date, effective) or \
           is_ticker_done(state, run_date, ticker):
            logger.info(
                f"[{ticker}] already done today -- skipping ({i}/{len(tickers)})"
            )
            skipped += 1
            continue

        logger.info(f"Processing {ticker}  ({i}/{len(tickers)})...")

        result, used_ticker = fetch_with_smart_retry(
            ticker, fixes, state, run_date, logger
        )

        if result is None:
            failed += 1
            continue

        if used_ticker != ticker:
            corrected += 1

        try:
            append_prediction_row(
                ticker       = used_ticker,
                run_date     = run_date,
                actual       = result["actual"],
                predicted    = result["predicted"],
                price_feats  = result["price_feats"],
                factor_feats = result["factor_feats"],
            )
            # Mark both original and corrected symbol as done
            mark_ticker_done(state, run_date, ticker)
            if used_ticker != ticker:
                mark_ticker_done(state, run_date, used_ticker)
            success += 1

        except Exception as exc:
            logger.error(f"[{used_ticker}] Excel write error: {exc}")
            failed += 1

    logger.info(
        f"=== Day complete  {run_date} ===  "
        f"success={success}  auto-corrected={corrected}  "
        f"skipped={skipped}  failed={failed}"
    )

    if fixes:
        logger.info("-- Current ticker fix map " + "-" * 44)
        for old, new in fixes.items():
            logger.info(f"   {old:30s} -> {new}")
        logger.info("-" * 70)


# =============================================================================
# SCHEDULER
# =============================================================================

def _next_run_time() -> datetime:
    now    = datetime.now()
    target = now.replace(hour=RUN_HOUR, minute=RUN_MINUTE, second=0, microsecond=0)
    if now >= target:
        target += timedelta(days=1)
    return target


def run_scheduler(tickers: list[str], resume: bool = False) -> None:
    state = load_state()

    if not resume or state["start_date"] is None:
        state["start_date"]     = datetime.now().strftime("%Y-%m-%d")
        state["days_completed"] = 0
        save_state(state)

    run_date = datetime.now().strftime("%Y-%m-%d")
    logger   = setup_logging(run_date)
    start_dt = datetime.strptime(state["start_date"], "%Y-%m-%d")
    end_dt   = start_dt + timedelta(days=TOTAL_DAYS)

    logger.info(
        f"Automation start: {state['start_date']}  |  "
        f"End: {end_dt.strftime('%Y-%m-%d')}  |  "
        f"Days completed so far: {state['days_completed']}"
    )

    while datetime.now() < end_dt:
        next_run = _next_run_time()
        wait_sec = (next_run - datetime.now()).total_seconds()
        logger.info(
            f"Next run scheduled at {next_run.strftime('%Y-%m-%d %H:%M:%S')} "
            f"(~{wait_sec / 3600:.1f} h away)"
        )
        time.sleep(max(0, wait_sec))

        today  = datetime.now().strftime("%Y-%m-%d")
        logger = setup_logging(today)

        run_daily(tickers, state, logger)
        state["days_completed"] += 1
        save_state(state)

    logger.info("30-day automation complete. All workbooks saved to ./workbooks/")


# =============================================================================
# ENTRY POINT
# =============================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ticker-Teller NIFTY 100 Daily Automation (v2 -- Smart Retry)"
    )
    parser.add_argument("--once",   action="store_true",
                        help="Run once immediately (no scheduler)")
    parser.add_argument("--resume", action="store_true",
                        help="Resume from last checkpoint")
    parser.add_argument("--tickers", nargs="+", metavar="TICKER",
                        help="Override ticker list (e.g. RELIANCE.NS TCS.NS)")
    args = parser.parse_args()

    tickers  = args.tickers if args.tickers else NIFTY100_TICKERS
    run_date = datetime.now().strftime("%Y-%m-%d")
    logger   = setup_logging(run_date)

    logger.info(
        f"Ticker-Teller NIFTY 100 Automation v2 | tickers={len(tickers)}"
    )

    if args.once:
        state = load_state()
        run_daily(tickers, state, logger)
    else:
        run_scheduler(tickers, resume=args.resume)


if __name__ == "__main__":
    main()
