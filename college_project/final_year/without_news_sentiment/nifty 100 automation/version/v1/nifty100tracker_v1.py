"""
nifty100_tracker.py — Ticker-Teller Daily Automation
=====================================================
Runs daily predictions for all NIFTY 100 companies and records results
into individual Excel workbooks with conditional formatting.

Usage:
    python nifty100_tracker.py               # Run full 30-day automation
    python nifty100_tracker.py --once        # Run a single day immediately
    python nifty100_tracker.py --resume      # Resume from last checkpoint
    python nifty100_tracker.py --tickers RELIANCE.NS TCS.NS  # Run specific tickers

Directory layout (auto-created):
    workbooks/   ← one .xlsx per company
    logs/        ← daily + master log files
    state/       ← checkpoint JSON for resumability
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import torch
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from sklearn.preprocessing import MinMaxScaler

# ── Path bootstrap so src/ imports work from any CWD ──────────────────────────
ROOT_DIR = Path(__file__).parent
sys.path.insert(0, str(ROOT_DIR))

from src.data import (
    FACTOR_FEATURE_COLS, FEATURE_COLS, N_FACTOR_FEATURES,
    N_PRICE_FEATURES, PRICE_FEATURE_COLS,
    add_quant_features, add_technical_features,
    build_sequences, fetch_data, fetch_market_data,
    returns_to_prices, split_data,
)
from src.model import (
    make_loaders, monte_carlo_forecast, predict_test_set, train_model,
)

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

WORKBOOK_DIR = ROOT_DIR / "workbooks"
LOG_DIR      = ROOT_DIR / "logs"
STATE_DIR    = ROOT_DIR / "state"

TOTAL_DAYS     = 30       # automation runs for this many calendar days
RUN_HOUR       = 18       # 6 PM IST — after NSE market close (15:30 IST)
RUN_MINUTE     = 0
RETRY_ATTEMPTS = 3        # per-ticker retry on transient failure
RETRY_DELAY    = 60       # seconds between retries

# Model hyper-parameters (mirrors Ticker-Teller v5 defaults)
MODEL_CFG = dict(
    period       = 730,   # historical days
    seq_length   = 60,
    forecast_days= 1,     # we only need next-day forecast
    epochs       = 50,
    patience     = 10,
    hidden_size  = 128,
    cnn_channels = 64,
    num_heads    = 4,
    dropout      = 0.2,
)

# ── Excel visual theme ─────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")   # dark navy
SUB_FILL     = PatternFill("solid", fgColor="0F2744")   # deeper navy for sub-headers
HEADER_FONT  = Font(bold=True, color="E2E8F0", name="Arial", size=10)
SUB_FONT     = Font(bold=True, color="94A3B8", name="Arial", size=9)
DATA_FONT    = Font(name="Arial", size=9)
MONO_FONT    = Font(name="Courier New", size=8, color="94A3B8")
THIN_BORDER  = Border(
    left   = Side(style="thin", color="1E3A5F"),
    right  = Side(style="thin", color="1E3A5F"),
    top    = Side(style="thin", color="1E3A5F"),
    bottom = Side(style="thin", color="1E3A5F"),
)
GREEN_FILL = PatternFill("solid", fgColor="052E16")     # for positive error
RED_FILL   = PatternFill("solid", fgColor="450A0A")     # for negative error

# ═══════════════════════════════════════════════════════════════════════════════
# NIFTY 100 TICKERS  (Yahoo Finance format — suffix .NS)
# ═══════════════════════════════════════════════════════════════════════════════

NIFTY100_TICKERS: list[str] = [
    "RELIANCE.NS","TCS.NS","HDFCBANK.NS","BHARTIARTL.NS","ICICIBANK.NS",
    "INFY.NS","SBIN.NS","HINDUNILVR.NS","ITC.NS","LT.NS",
    "BAJFINANCE.NS","HCLTECH.NS","MARUTI.NS","SUNPHARMA.NS","ADANIENT.NS",
    "KOTAKBANK.NS","TITAN.NS","ONGC.NS","NTPC.NS","POWERGRID.NS",
    "ULTRACEMCO.NS","WIPRO.NS","BAJAJFINSV.NS","ADANIPORTS.NS","NESTLEIND.NS",
    "JSWSTEEL.NS","TATASTEEL.NS","HINDALCO.NS","DIVISLAB.NS","DRREDDY.NS",
    "CIPLA.NS","ASIANPAINT.NS","COALINDIA.NS","BPCL.NS","GRASIM.NS",
    "EICHERMOT.NS","HEROMOTOCO.NS","APOLLOHOSP.NS","BRITANNIA.NS","TATACONSUM.NS",
    "M&M.NS","SBILIFE.NS","HDFCLIFE.NS","SHRIRAMFIN.NS","BAJAJ-AUTO.NS",
    "INDUSINDBK.NS","TECHM.NS","TMCV.NS", "TMPV.NS", "AXISBANK.NS","LTF.NS",
    "LTIM.NS","PIDILITIND.NS","DMART.NS","HAVELLS.NS","SIEMENS.NS",
    "MUTHOOTFIN.NS","BEL.NS","GODREJCP.NS","BERGEPAINT.NS","CHOLAFIN.NS",
    "TRENT.NS","COLPAL.NS","AMBUJACEM.NS","TORNTPHARM.NS","DABUR.NS",
    "SRF.NS","MARICO.NS","BANKBARODA.NS","NAUKRI.NS","ETERNAL.NS",
    "ICICIPRULI.NS","JIOFIN.NS","INDIGO.NS","PNB.NS","ADANIGREEN.NS",
    "ADANIENSOL.NS","ADANIPOWER.NS","ABB.NS","MOTHERSON.NS","PGHH.NS",
    "MAXHEALTH.NS","OBEROIRLTY.NS","MFSL.NS","PERSISTENT.NS","POLYCAB.NS",
    "CANBK.NS","INDIANB.NS","PETRONET.NS","GAIL.NS","IOC.NS",
    "RECLTD.NS","PFC.NS","NHPC.NS","IRFC.NS","HUDCO.NS",
    "LINDEINDIA.NS","CGPOWER.NS","PHOENIXLTD.NS","TIINDIA.NS","NYKAA.NS",
]

# ═══════════════════════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

def setup_logging(run_date: str) -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f"tracker_{run_date}.log"
    master   = LOG_DIR / "master.log"

    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    logger = logging.getLogger("nifty100_tracker")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    mh = logging.FileHandler(master, encoding="utf-8")
    mh.setLevel(logging.INFO)
    mh.setFormatter(fmt)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(mh)
    logger.addHandler(ch)
    return logger


# ═══════════════════════════════════════════════════════════════════════════════
# STATE / CHECKPOINT
# ═══════════════════════════════════════════════════════════════════════════════

STATE_FILE = STATE_DIR / "checkpoint.json"

def load_state() -> dict:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    if STATE_FILE.exists():
        with open(STATE_FILE) as f:
            return json.load(f)
    return {"days_completed": 0, "start_date": None, "daily_records": {}}


def save_state(state: dict) -> None:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2, default=str)


def mark_ticker_done(state: dict, run_date: str, ticker: str) -> None:
    state.setdefault("daily_records", {}).setdefault(run_date, [])
    if ticker not in state["daily_records"][run_date]:
        state["daily_records"][run_date].append(ticker)
    save_state(state)


def is_ticker_done(state: dict, run_date: str, ticker: str) -> bool:
    return ticker in state.get("daily_records", {}).get(run_date, [])


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL WORKBOOK HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

# Column layout:
# A: Date | B: Ticker | C: Price Features | D: Factor Features
# E: Actual Price | F: Predicted Price | G: Error (%)

COL_DATE     = 1   # A
COL_TICKER   = 2   # B
COL_PRICE_F  = 3   # C  — price branch features (16)
COL_FACTOR_F = 4   # D  — quant factor features (6)
COL_ACTUAL   = 5   # E
COL_PRED     = 6   # F
COL_ERROR    = 7   # G

HEADER_ROW   = 1   # top-level merged header
SUB_ROW      = 2   # sub-column labels
DATA_START   = 3   # first data row


def _safe_filename(ticker: str) -> str:
    return ticker.replace(".", "_").replace("/", "_").replace("&", "AND")


def _workbook_path(ticker: str) -> Path:
    WORKBOOK_DIR.mkdir(parents=True, exist_ok=True)
    return WORKBOOK_DIR / f"{_safe_filename(ticker)}.xlsx"


def _apply_header_border(cell, fill=None, font=None, align_center=True) -> None:
    if fill:  cell.fill   = fill
    if font:  cell.font   = font
    cell.border = THIN_BORDER
    if align_center:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def create_workbook(ticker: str) -> None:
    """Create a fresh workbook with professional headers for a ticker."""
    wb   = Workbook()
    ws   = wb.active
    ws.title = "Predictions"

    company = ticker.replace(".NS", "").replace("-", " ")
    ws.sheet_properties.tabColor = "1E3A5F"

    # ── Row 1: top-level merged group headers ──────────────────────────────────
    # Date
    ws.merge_cells(start_row=1, start_column=COL_DATE, end_row=2, end_column=COL_DATE)
    c = ws.cell(1, COL_DATE, "Date")
    _apply_header_border(c, HEADER_FILL, HEADER_FONT)

    # Ticker
    ws.merge_cells(start_row=1, start_column=COL_TICKER, end_row=2, end_column=COL_TICKER)
    c = ws.cell(1, COL_TICKER, "Ticker")
    _apply_header_border(c, HEADER_FILL, HEADER_FONT)

    # Model Input (spans Price Features + Factor Features)
    ws.merge_cells(start_row=1, start_column=COL_PRICE_F,
                   end_row=1, end_column=COL_FACTOR_F)
    c = ws.cell(1, COL_PRICE_F, "Model Input")
    _apply_header_border(c, HEADER_FILL, HEADER_FONT)

    # Sub-headers for Model Input
    c = ws.cell(SUB_ROW, COL_PRICE_F,
                f"Price Features ({N_PRICE_FEATURES})\n"
                + " · ".join(PRICE_FEATURE_COLS))
    _apply_header_border(c, SUB_FILL, SUB_FONT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    c = ws.cell(SUB_ROW, COL_FACTOR_F,
                f"Quant Factor Features ({N_FACTOR_FEATURES})\n"
                + " · ".join(FACTOR_FEATURE_COLS))
    _apply_header_border(c, SUB_FILL, SUB_FONT)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Actual Price
    ws.merge_cells(start_row=1, start_column=COL_ACTUAL,
                   end_row=2, end_column=COL_ACTUAL)
    c = ws.cell(1, COL_ACTUAL, "Actual Price\n(₹)")
    _apply_header_border(c, HEADER_FILL, HEADER_FONT)

    # Predicted Price
    ws.merge_cells(start_row=1, start_column=COL_PRED,
                   end_row=2, end_column=COL_PRED)
    c = ws.cell(1, COL_PRED, "Predicted Price\n(₹)")
    _apply_header_border(c, HEADER_FILL, HEADER_FONT)

    # Error (%)
    ws.merge_cells(start_row=1, start_column=COL_ERROR,
                   end_row=2, end_column=COL_ERROR)
    c = ws.cell(1, COL_ERROR,
                "Error (%)\n|((Actual−Pred)/Actual)×100|")
    _apply_header_border(c, HEADER_FILL, HEADER_FONT)

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(COL_DATE)].width     = 14
    ws.column_dimensions[get_column_letter(COL_TICKER)].width   = 14
    ws.column_dimensions[get_column_letter(COL_PRICE_F)].width  = 60
    ws.column_dimensions[get_column_letter(COL_FACTOR_F)].width = 50
    ws.column_dimensions[get_column_letter(COL_ACTUAL)].width   = 16
    ws.column_dimensions[get_column_letter(COL_PRED)].width     = 16
    ws.column_dimensions[get_column_letter(COL_ERROR)].width    = 22

    # ── Row heights ────────────────────────────────────────────────────────────
    ws.row_dimensions[HEADER_ROW].height = 28
    ws.row_dimensions[SUB_ROW].height    = 70

    # ── Freeze panes at first data row ────────────────────────────────────────
    ws.freeze_panes = ws.cell(DATA_START, 1)

    # ── Title sheet ───────────────────────────────────────────────────────────
    ts = wb.create_sheet("Info")
    ts["A1"] = f"Ticker-Teller  —  {company}"
    ts["A1"].font = Font(bold=True, size=14, color="38BDF8", name="Arial")
    ts["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ts["A2"].font = Font(size=9, color="94A3B8", name="Arial")
    ts["A3"] = "Model: CNN → BiLSTM → Multi-Head Attention"
    ts["A4"] = f"Price features  ({N_PRICE_FEATURES}): " + ", ".join(PRICE_FEATURE_COLS)
    ts["A5"] = f"Factor features ({N_FACTOR_FEATURES}): " + ", ".join(FACTOR_FEATURE_COLS)
    ts["A6"] = "Target: log return = log(Close[t] / Close[t-1])"
    ts["A7"] = "Error formula: ABS(((Actual - Predicted) / Actual) * 100)"
    ts["A8"] = "Conditional format — GREEN cell: error > 0 (model underestimated); RED cell: error < 0 (overestimated)"
    for r in range(1, 9):
        ts.cell(r, 1).alignment = Alignment(wrap_text=True)
    ts.column_dimensions["A"].width = 90

    wb.save(_workbook_path(ticker))


def append_prediction_row(
    ticker:      str,
    run_date:    str,
    actual:      float,
    predicted:   float,
    price_feats: str,
    factor_feats:str,
) -> None:
    """Append one data row to the ticker's workbook; apply conditional formatting."""
    path = _workbook_path(ticker)
    if not path.exists():
        create_workbook(ticker)

    wb = load_workbook(path)
    ws = wb["Predictions"]

    next_row = ws.max_row + 1
    if next_row < DATA_START:
        next_row = DATA_START

    # Signed error (used for CF direction) stored in a helper col but hidden
    # Absolute error shown in col G using ABS formula; sign drives CF via
    # conditional formatting rule that checks the signed raw value.
    # We store signed raw in column H (hidden) so CF can reference it.
    COL_SIGNED = 8  # H — hidden raw signed error

    # Write data cells
    ws.cell(next_row, COL_DATE,   run_date).number_format   = "YYYY-MM-DD"
    ws.cell(next_row, COL_TICKER, ticker)
    ws.cell(next_row, COL_PRICE_F,  price_feats)
    ws.cell(next_row, COL_FACTOR_F, factor_feats)

    e_col = get_column_letter(COL_ACTUAL)
    f_col = get_column_letter(COL_PRED)
    g_col = get_column_letter(COL_ERROR)
    h_col = get_column_letter(COL_SIGNED)

    ws.cell(next_row, COL_ACTUAL, actual).number_format    = '₹#,##0.00'
    ws.cell(next_row, COL_PRED,   predicted).number_format = '₹#,##0.00'

    # Absolute error formula (display)
    ws.cell(next_row, COL_ERROR,
            f'=ABS((({e_col}{next_row}-{f_col}{next_row})/{e_col}{next_row})*100)'
    ).number_format = '0.00"%"'

    # Signed error (hidden — drives conditional formatting)
    ws.cell(next_row, COL_SIGNED,
            f'=(({e_col}{next_row}-{f_col}{next_row})/{e_col}{next_row})*100'
    ).number_format = '0.00"%"'

    # ── Style data row ─────────────────────────────────────────────────────────
    for col in range(1, COL_SIGNED + 1):
        cell = ws.cell(next_row, col)
        cell.border = THIN_BORDER
        if col in (COL_PRICE_F, COL_FACTOR_F):
            cell.font      = MONO_FONT
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)
        else:
            cell.font      = DATA_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[next_row].height = 14

    # ── Conditional formatting on col G (ABS error cell) ──────────────────────
    # Rule is on the SIGNED column (H): green when H > 0, red when H < 0
    error_range = f"{g_col}{DATA_START}:{g_col}{next_row}"
    ws.conditional_formatting._cf_rules = {}  # clear old rules

    ws.conditional_formatting.add(
        error_range,
        CellIsRule(
            operator     = "greaterThan",
            formula      = [f"0"],
            stopIfTrue   = False,
            fill         = PatternFill("solid", fgColor="052E16"),
            font         = Font(color="4ADE80", bold=True, name="Arial", size=9),
        ),
    )
    # The CF "greaterThan 0" on the displayed ABS value is always True,
    # so instead we apply CF based on the SIGNED column H:
    signed_range = f"{g_col}{DATA_START}:{g_col}{next_row}"
    ws.conditional_formatting._cf_rules = {}

    # Green: actual > predicted  (model undershot → positive signed error)
    ws.conditional_formatting.add(
        signed_range,
        CellIsRule(
            operator  = "greaterThan",
            formula   = [f"({h_col}{DATA_START})"],
            stopIfTrue= False,
            fill      = PatternFill("solid", fgColor="052E16"),
            font      = Font(color="4ADE80", bold=True, name="Arial", size=9),
        ),
    )

    # Apply green/red directly based on the computed signed error value
    raw_signed = ((actual - predicted) / actual * 100) if actual != 0 else 0
    fill_color = PatternFill("solid", fgColor="052E16") if raw_signed >= 0 \
                 else PatternFill("solid", fgColor="450A0A")
    font_color = "4ADE80" if raw_signed >= 0 else "FCA5A5"

    ws.cell(next_row, COL_ERROR).fill = fill_color
    ws.cell(next_row, COL_ERROR).font = Font(
        color=font_color, bold=True, name="Arial", size=9
    )

    # Hide the signed helper column
    ws.column_dimensions[h_col].hidden = True

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE SNAPSHOT HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _feature_snapshot(df: pd.DataFrame, feats: list[str]) -> str:
    """Return the last row's feature values as a compact readable string."""
    last = df[feats].dropna().iloc[-1]
    parts = [f"{k}={v:.4f}" for k, v in last.items()]
    return " | ".join(parts)


# ═══════════════════════════════════════════════════════════════════════════════
# CORE PREDICTION LOGIC  (wraps Ticker-Teller v5 pipeline)
# ═══════════════════════════════════════════════════════════════════════════════

def run_prediction(ticker: str, logger: logging.Logger) -> Optional[dict]:
    """
    Full pipeline for one ticker.
    Returns dict with keys: actual, predicted, price_feats_str, factor_feats_str
    or None on failure.
    """
    cfg = MODEL_CFG
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    logger.debug(f"[{ticker}] Fetching OHLCV…")
    df_raw = fetch_data(ticker, cfg["period"])
    if df_raw is None or df_raw.empty:
        logger.warning(f"[{ticker}] No OHLCV data — skipping.")
        return None

    logger.debug(f"[{ticker}] Fetching SPY market data…")
    market_df = fetch_market_data(cfg["period"])
    if market_df is None or market_df.empty:
        logger.warning(f"[{ticker}] No SPY data — using Close proxy.")
        market_df = df_raw[["Close"]].copy()

    # Feature engineering
    df = add_technical_features(df_raw)
    df = add_quant_features(df, market_df)
    df = df.dropna(subset=FEATURE_COLS)[FEATURE_COLS]

    if len(df) < cfg["seq_length"] + 10:
        logger.warning(f"[{ticker}] Insufficient rows ({len(df)}) — skipping.")
        return None

    # Feature snapshots (last available bar before forecast)
    price_feats_str  = _feature_snapshot(df, PRICE_FEATURE_COLS)
    factor_feats_str = _feature_snapshot(df, FACTOR_FEATURE_COLS)

    # Scale
    scaler      = MinMaxScaler(feature_range=(0, 1))
    scaled_data = scaler.fit_transform(df)
    raw_close   = df["Close"].values
    X, y        = build_sequences(scaled_data, raw_close, cfg["seq_length"])

    (X_train, y_train), (X_val, y_val), (X_test, y_test), split, val_split = \
        split_data(X, y)

    train_loader, val_loader, test_loader = make_loaders(
        X_train, y_train, X_val, y_val, X_test, y_test
    )

    # Redirect Streamlit progress calls (not available in CLI)
    import unittest.mock as mock
    dummy = mock.MagicMock()
    with mock.patch("streamlit.progress", return_value=dummy), \
         mock.patch("streamlit.empty",    return_value=dummy):

        logger.debug(f"[{ticker}] Training model…")
        model, _ = train_model(
            train_loader, val_loader,
            input_size   = len(FEATURE_COLS),
            epochs       = cfg["epochs"],
            patience     = cfg["patience"],
            device       = device,
            hidden_size  = cfg["hidden_size"],
            dropout      = cfg["dropout"],
            cnn_channels = cfg["cnn_channels"],
            num_heads    = cfg["num_heads"],
        )

    # Monte Carlo 1-day forecast
    last_close = float(df_raw["Close"].iloc[-1])
    mc_raw = monte_carlo_forecast(
        model,
        scaled_data,
        raw_close_last = last_close,
        scaler         = scaler,
        n_features     = len(FEATURE_COLS),
        seq_length     = cfg["seq_length"],
        forecast_days  = 1,
        device         = device,
    )
    predicted_price = float(mc_raw[0]["mean"])
    actual_price    = last_close   # "actual" = last known close at run time

    logger.info(
        f"[{ticker}] actual={actual_price:.2f}  predicted={predicted_price:.2f}  "
        f"error={abs((actual_price-predicted_price)/actual_price*100):.2f}%"
    )
    return dict(
        actual        = actual_price,
        predicted     = predicted_price,
        price_feats   = price_feats_str,
        factor_feats  = factor_feats_str,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# DAILY RUN  (one pass over all tickers)
# ═══════════════════════════════════════════════════════════════════════════════

def run_daily(tickers: list[str], state: dict, logger: logging.Logger) -> None:
    run_date = datetime.now().strftime("%Y-%m-%d")
    logger.info(f"═══ Daily run  {run_date} ═══  ({len(tickers)} tickers)")

    success, skipped, failed = 0, 0, 0

    for i, ticker in enumerate(tickers, 1):
        if is_ticker_done(state, run_date, ticker):
            logger.info(f"[{ticker}] already done today — skipping ({i}/{len(tickers)})")
            skipped += 1
            continue

        logger.info(f"Processing {ticker}  ({i}/{len(tickers)})…")

        result = None
        for attempt in range(1, RETRY_ATTEMPTS + 1):
            try:
                result = run_prediction(ticker, logger)
                break
            except Exception as exc:
                logger.warning(f"[{ticker}] attempt {attempt} failed: {exc}")
                if attempt < RETRY_ATTEMPTS:
                    time.sleep(RETRY_DELAY)

        if result is None:
            logger.error(f"[{ticker}] all attempts failed — skipping.")
            failed += 1
            continue

        try:
            append_prediction_row(
                ticker       = ticker,
                run_date     = run_date,
                actual       = result["actual"],
                predicted    = result["predicted"],
                price_feats  = result["price_feats"],
                factor_feats = result["factor_feats"],
            )
            mark_ticker_done(state, run_date, ticker)
            success += 1
        except Exception as exc:
            logger.error(f"[{ticker}] Excel write error: {exc}")
            failed += 1

    logger.info(
        f"═══ Day complete  {run_date} ═══  "
        f"success={success}  skipped={skipped}  failed={failed}"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# SCHEDULER  (blocks until 30-day run completes)
# ═══════════════════════════════════════════════════════════════════════════════

def _next_run_time() -> datetime:
    now = datetime.now()
    target = now.replace(hour=RUN_HOUR, minute=RUN_MINUTE, second=0, microsecond=0)
    if now >= target:
        target += timedelta(days=1)
    return target


def run_scheduler(tickers: list[str], resume: bool = False) -> None:
    state = load_state()

    if not resume or state["start_date"] is None:
        state["start_date"]     = datetime.now().strftime("%Y-%m-%d")
        state["days_completed"] = 0
        save_state(state)

    run_date  = datetime.now().strftime("%Y-%m-%d")
    logger    = setup_logging(run_date)
    start_dt  = datetime.strptime(state["start_date"], "%Y-%m-%d")
    end_dt    = start_dt + timedelta(days=TOTAL_DAYS)

    logger.info(f"Automation start: {state['start_date']}  |  "
                f"End: {end_dt.strftime('%Y-%m-%d')}  |  "
                f"Days completed so far: {state['days_completed']}")

    while datetime.now() < end_dt:
        next_run = _next_run_time()
        wait_sec = (next_run - datetime.now()).total_seconds()
        logger.info(f"Next run scheduled at {next_run.strftime('%Y-%m-%d %H:%M:%S')} "
                    f"(~{wait_sec/3600:.1f} h away)")
        time.sleep(max(0, wait_sec))

        # Re-setup logging for the new day
        today  = datetime.now().strftime("%Y-%m-%d")
        logger = setup_logging(today)

        run_daily(tickers, state, logger)
        state["days_completed"] += 1
        save_state(state)

    logger.info("30-day automation complete. All workbooks saved to ./workbooks/")


# ═══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ticker-Teller NIFTY 100 Daily Automation"
    )
    parser.add_argument("--once",    action="store_true",
                        help="Run once immediately (no scheduler)")
    parser.add_argument("--resume",  action="store_true",
                        help="Resume from last checkpoint")
    parser.add_argument("--tickers", nargs="+", metavar="TICKER",
                        help="Override ticker list (e.g. RELIANCE.NS TCS.NS)")
    args = parser.parse_args()

    tickers = args.tickers if args.tickers else NIFTY100_TICKERS
    run_date = datetime.now().strftime("%Y-%m-%d")
    logger   = setup_logging(run_date)

    logger.info(f"Ticker-Teller NIFTY 100 Automation | tickers={len(tickers)}")

    if args.once:
        state = load_state()
        run_daily(tickers, state, logger)
    else:
        run_scheduler(tickers, resume=args.resume)


if __name__ == "__main__":
    main()
